"""临床诊断映射 Excel 增量导入服务（101号 §4.2）。

支持两行表头、OCR 暂存、校验、审核和正式合并。
不复用现有八代码集全量删除重建脚本。
"""
from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Optional

from sqlalchemy import select, func
from sqlalchemy.orm import Session

from ..models.dict_medical import (
    DictMedicalCodeItem,
    DictMedicalCodeMapping,
    DictMedicalCodeSet,
    DictMedicalImportRun,
)
from ..models.dict_medical_push import DictMedicalImportRow

MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_ROWS = 5000
EXPECTED_SHEET = "诊断字典映射"

COLUMN_ALIASES = {
    "字典属性": "dict_attribute",
    "院内疾病编码": "hospital_code",
    "院内疾病名称": "hospital_name",
    "国家临床版2.0编码": "national_clinical_code",
    "国家临床版2.0名称": "national_clinical_name",
    "国家医保版2.0编码": "insurance_code",
    "国家医保版2.0名称": "insurance_name",
}

CODE_SET_HOSPITAL = "hospital_diagnosis"
CODE_SET_NATIONAL_CLINICAL = "national_clinical_2"
CODE_SET_INSURANCE = "national_insurance_2"
CATEGORY = "diagnosis"


def _normalize_text(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", s)
    s = s.replace("\u3000", " ").strip()
    return s if s else None


def _row_hash(values: dict[str, Optional[str]]) -> str:
    raw = "|".join(values.get(k) or "" for k in sorted(values.keys()))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:32]


def parse_diagnosis_mapping_excel(file_bytes: bytes, file_name: str) -> dict[str, Any]:
    """解析两行表头的诊断映射 Excel，返回结构化结果。"""
    import openpyxl

    if len(file_bytes) > MAX_FILE_SIZE:
        return {"error": f"文件超过 {MAX_FILE_SIZE // 1024 // 1024}MB 限制"}

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    target_sheet = None
    for name in sheet_names:
        if EXPECTED_SHEET in name or "诊断" in name:
            target_sheet = name
            break
    if not target_sheet:
        target_sheet = sheet_names[0] if sheet_names else None
    if not target_sheet:
        return {"error": "Excel 无工作表"}

    ws = wb[target_sheet]
    rows_iter = ws.iter_rows(values_only=True)

    header_row1 = next(rows_iter, None)
    header_row2 = next(rows_iter, None)
    if header_row1 is None:
        return {"error": "Excel 为空"}

    col_map: dict[int, str] = {}
    for row in (header_row1, header_row2):
        if row is None:
            continue
        for idx, cell in enumerate(row):
            cell_str = _normalize_text(cell)
            if cell_str:
                for alias, field in COLUMN_ALIASES.items():
                    if alias in cell_str and field not in col_map.values():
                        col_map[idx] = field
                        break

    if "hospital_code" not in col_map.values():
        return {"error": "未识别到院内疾病编码列，请检查表头"}

    data_rows: list[dict[str, Any]] = []
    for row_no, row in enumerate(rows_iter, start=3):
        values: dict[str, Optional[str]] = {}
        for idx, field in col_map.items():
            values[field] = _normalize_text(row[idx]) if idx < len(row) else None
        if not values.get("hospital_code"):
            continue
        if len(data_rows) >= MAX_ROWS:
            return {"error": f"数据行超过 {MAX_ROWS} 行限制"}
        data_rows.append({"row_no": row_no, "values": values})

    wb.close()
    file_sha = hashlib.sha256(file_bytes).hexdigest()
    return {
        "sheet": target_sheet,
        "file_sha256": file_sha,
        "row_count": len(data_rows),
        "rows": data_rows,
        "col_map": {v: k for k, v in col_map.items()},
    }


def validate_row(values: dict[str, Optional[str]], existing_codes: set[str]) -> dict[str, Any]:
    """校验单行数据，返回 validation_status 和 errors。"""
    errors: list[str] = []
    hospital_code = values.get("hospital_code")
    hospital_name = values.get("hospital_name")
    nc_code = values.get("national_clinical_code")
    ins_code = values.get("insurance_code")
    ins_name = values.get("insurance_name")

    if not hospital_code:
        errors.append("院内编码为空")
    if not hospital_name:
        errors.append("院内名称为空")

    insurance_status = "valid"
    if not ins_code:
        insurance_status = "empty"
    elif ins_name is None or ins_name == "":
        insurance_status = "grey"
    elif not re.match(r"^[A-Z]\d{2}", ins_code or ""):
        insurance_status = "invalid"
        errors.append(f"医保编码格式异常: {ins_code}")

    validation_status = "valid"
    if errors:
        validation_status = "error"
    elif insurance_status == "grey":
        validation_status = "warning"

    diff_type = "new"
    if hospital_code and hospital_code in existing_codes:
        diff_type = "exact_match"

    return {
        "validation_status": validation_status,
        "validation_errors": errors if errors else None,
        "insurance_mapping_status": insurance_status,
        "diff_type": diff_type,
    }


def create_import_run(
    db: Session,
    *,
    file_name: str,
    file_sha256: str,
    sheet: str,
    row_count: int,
    operator: str,
) -> DictMedicalImportRun:
    """创建导入批次（幂等：同 SHA 返回既有批次）。"""
    existing = db.scalar(
        select(DictMedicalImportRun).where(
            DictMedicalImportRun.diagnosis_sha256 == file_sha256
        )
    )
    if existing:
        return existing

    batch_code = f"diag-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{file_sha256[:8]}"
    run = DictMedicalImportRun(
        batch_code=batch_code,
        source_dir="upload",
        diagnosis_file_name=file_name,
        diagnosis_sha256=file_sha256,
        status="staged",
        mode="incremental",
        operator=operator,
        stats={"row_count": row_count, "sheet": sheet},
    )
    db.add(run)
    db.flush()
    return run


def stage_rows(
    db: Session,
    run: DictMedicalImportRun,
    rows: list[dict[str, Any]],
    file_name: str,
    file_sha256: str,
    sheet: str,
) -> int:
    """将解析行写入暂存表。"""
    existing_codes = set(
        db.scalars(
            select(DictMedicalCodeItem.item_code).where(
                DictMedicalCodeItem.code_set_code == CODE_SET_HOSPITAL,
                DictMedicalCodeItem.category_code == CATEGORY,
            )
        ).all()
    )

    seen_codes: dict[str, int] = {}
    staged = 0
    for row_data in rows:
        values = row_data["values"]
        row_no = row_data["row_no"]
        hospital_code = values.get("hospital_code") or ""

        diff_override = None
        if hospital_code in seen_codes:
            diff_override = "duplicate_in_file"
        seen_codes[hospital_code] = row_no

        result = validate_row(values, existing_codes)
        if diff_override:
            result["diff_type"] = diff_override
            result["validation_status"] = "warning"
            errs = result.get("validation_errors") or []
            errs.append(f"本批重复（首次出现在行 {seen_codes[hospital_code]}）")
            result["validation_errors"] = errs

        row_obj = DictMedicalImportRow(
            import_run_id=run.id,
            source_file_name=file_name,
            source_file_sha256=file_sha256,
            source_sheet=sheet,
            source_row_no=row_no,
            row_hash=_row_hash(values),
            raw_dict_attribute=values.get("dict_attribute"),
            raw_hospital_code=values.get("hospital_code"),
            raw_hospital_name=values.get("hospital_name"),
            raw_national_clinical_code=values.get("national_clinical_code"),
            raw_national_clinical_name=values.get("national_clinical_name"),
            raw_insurance_code=values.get("insurance_code"),
            raw_insurance_name=values.get("insurance_name"),
            norm_dict_attribute=values.get("dict_attribute"),
            norm_hospital_code=values.get("hospital_code"),
            norm_hospital_name=values.get("hospital_name"),
            norm_national_clinical_code=values.get("national_clinical_code"),
            norm_national_clinical_name=values.get("national_clinical_name"),
            norm_insurance_code=values.get("insurance_code"),
            norm_insurance_name=values.get("insurance_name"),
            insurance_mapping_status=result["insurance_mapping_status"],
            validation_status=result["validation_status"],
            validation_errors=result["validation_errors"],
            diff_type=result["diff_type"],
            review_status="pending",
        )
        db.add(row_obj)
        staged += 1

    db.commit()
    return staged


def merge_approved_rows(db: Session, run: DictMedicalImportRun, merged_by: str) -> dict[str, int]:
    """将已审核通过的暂存行增量合并到正式字典。不删除历史数据。"""
    approved = db.scalars(
        select(DictMedicalImportRow).where(
            DictMedicalImportRow.import_run_id == run.id,
            DictMedicalImportRow.review_status == "approved",
            DictMedicalImportRow.merged_at.is_(None),
        )
    ).all()

    stats = {"merged": 0, "skipped_exact": 0, "conflict": 0, "items_created": 0, "mappings_created": 0}
    now = datetime.now(timezone.utc)

    for row in approved:
        code = row.norm_hospital_code
        name = row.norm_hospital_name
        if not code or not name:
            stats["conflict"] += 1
            continue

        existing_item = db.scalar(
            select(DictMedicalCodeItem).where(
                DictMedicalCodeItem.code_set_code == CODE_SET_HOSPITAL,
                DictMedicalCodeItem.item_code == code,
            )
        )
        if existing_item:
            if existing_item.item_name_cn == name:
                stats["skipped_exact"] += 1
            else:
                stats["conflict"] += 1
                row.validation_status = "error"
                row.validation_errors = (row.validation_errors or []) + [f"名称冲突: 平台={existing_item.item_name_cn}, 导入={name}"]
                continue
        else:
            db.add(DictMedicalCodeItem(
                code_set_code=CODE_SET_HOSPITAL,
                item_code=code,
                item_name_cn=name,
                category_code=CATEGORY,
                status="active",
            ))
            stats["items_created"] += 1

        if row.norm_national_clinical_code:
            _ensure_mapping(db, code, CODE_SET_NATIONAL_CLINICAL, row.norm_national_clinical_code, row, stats)

        if row.norm_insurance_code and row.insurance_mapping_status == "valid":
            _ensure_mapping(db, code, CODE_SET_INSURANCE, row.norm_insurance_code, row, stats)

        row.merged_at = now
        row.merged_by = merged_by
        stats["merged"] += 1

    run.status = "merged"
    run.finished_at = now
    db.commit()
    return stats


def _ensure_mapping(db: Session, from_code: str, to_set: str, to_code: str, row: DictMedicalImportRow, stats: dict) -> None:
    existing = db.scalar(
        select(DictMedicalCodeMapping).where(
            DictMedicalCodeMapping.category_code == CATEGORY,
            DictMedicalCodeMapping.from_code_set == CODE_SET_HOSPITAL,
            DictMedicalCodeMapping.from_item_code == from_code,
            DictMedicalCodeMapping.to_code_set == to_set,
            DictMedicalCodeMapping.to_item_code == to_code,
        )
    )
    if not existing:
        db.add(DictMedicalCodeMapping(
            category_code=CATEGORY,
            from_code_set=CODE_SET_HOSPITAL,
            from_item_code=from_code,
            to_code_set=to_set,
            to_item_code=to_code,
            mapping_type="import",
            confidence="ocr_pending_review",
            review_status="approved",
            reviewer=row.reviewer,
            reviewed_at=row.reviewed_at,
        ))
        stats["mappings_created"] += 1